#!C:\pythonCode
# -*- coding: utf-8 -*-
# @Time : 2024/2/21 11:36
# @Author : Hank
# @File : excel_parse.py
# @Software: PyCharm

def parse_logs():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from parse import search

    workbook = openpyxl.load_workbook("data/user_logs.xlsx")
    sheet = workbook.active
    print(sheet.max_row)
    for idx in range(sheet.max_row):
        cell = sheet.cell(row=(idx + 2), column=7)
        cell2 = sheet.cell(row=(idx + 2), column=6)
        context = ""
        if cell.value is None:
            continue
        r = search("Hank\n\n{c1}\n\n{c2}\n\n{c3}\n\n{c4}\n\n{c5}", cell.value)
        # print(result1.evaluate_result.)
        if r is not None:
            if r['c1'] is not None and r['c1'] != "\n":
                context = context + r['c1']
                if r['c2'] is not None and r['c2'] != "\n":
                    context = context + "\n" + r['c2']
                    if r['c3'] is not None and r['c3'] != "\n":
                        context = context + "\n" + r['c3']
                        if r['c4'] is not None and r['c4'] != "\n":
                            context = context + "\n" + r['c4']
                            if r['c5'] is not None and r['c5'] != "\n":
                                context = context + "\n" + r['c5']
        else:
            r = search("Hank{c1}\n\n", cell.value)
            if r is not None:
                context = context + r['c1']
        cell_idx = get_column_letter(8) + str(idx + 2)
        print(cell2.value, context)
        sheet[cell_idx] = context
        print("----------------")
    workbook.save('result.xlsx')

if __name__ == '__main__':
    parse_logs()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/